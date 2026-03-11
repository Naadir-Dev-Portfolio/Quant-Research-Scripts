#!/usr/bin/env python3
"""
Fetch a stock's price history and CPI, then export to Excel with inflation-adjusted formulas.

Usage:
  python stock_inflation_to_excel.py --ticker AAPL --start 2000-01-01 --end 2025-08-13 --interval 1mo

Requires:
  pip install yfinance pandas pandas_datareader openpyxl python-dateutil

Notes:
  - CPI series: FRED "CPIAUCSL" (U.S. CPI, seasonally adjusted, 1982-84=100), monthly.
  - Stock prices are resampled to month-end to align with CPI by default when using 1mo interval.
  - Excel output includes formulas for CPI factor and real (inflation-adjusted) price.
"""

import argparse
from datetime import datetime
import sys
import pandas as pd

def main():
    parser = argparse.ArgumentParser(description="Export stock & CPI to Excel with inflation-adjusted formulas.")
    parser.add_argument("--ticker", default="AAPL", help="Ticker symbol, e.g., AAPL")
    parser.add_argument("--start", default="2000-01-01", help="Start date YYYY-MM-DD")
    parser.add_argument("--end", default=datetime.today().strftime("%Y-%m-%d"), help="End date YYYY-MM-DD")
    parser.add_argument("--interval", default="1mo", choices=["1d", "1wk", "1mo"], help="Sampling interval for prices")
    parser.add_argument("--outfile", default=None, help="Output Excel file (default: <TICKER>_inflation.xlsx)")
    args = parser.parse_args()

    try:
        import yfinance as yf
    except Exception as e:
        print("yfinance not installed. Run: pip install yfinance", file=sys.stderr)
        raise

    try:
        from pandas_datareader import data as pdr
    except Exception as e:
        print("pandas_datareader not installed. Run: pip install pandas_datareader", file=sys.stderr)
        raise

    # Download stock data
    ticker = args.ticker.upper()
    print(f"Downloading {ticker} price history {args.start} to {args.end} at {args.interval}...", file=sys.stderr)
    yf.pdr_override()
    # yfinance with pandas_datareader override allows DataReader("AAPL")? We'll stick with yf directly.
    hist = yf.download(ticker, start=args.start, end=args.end, interval=args.interval, auto_adjust=True, progress=False)
    if hist.empty:
        raise SystemExit(f"No price data returned for {ticker}. Check the symbol or date range.")
    # Use adjusted close if available; yfinance with auto_adjust=True puts it in 'Close'
    if "Adj Close" in hist.columns:
        price = hist[["Adj Close"]].rename(columns={"Adj Close": "AdjClose"})
    else:
        price = hist[["Close"]].rename(columns={"Close": "AdjClose"})

    # Normalize index to date (remove timezone) and ensure it's a DatetimeIndex
    price.index = pd.to_datetime(price.index).tz_localize(None)
    price = price[~price.index.duplicated(keep="last")]

    # For weekly/daily intervals, align CPI by month; we'll merge on month-end
    # Create a MonthEnd key for joining
    price["Month"] = price.index.to_period("M").dt.to_timestamp("M")

    # Download CPI from FRED
    print("Downloading CPIAUCSL from FRED...", file=sys.stderr)
    cpi = pdr.DataReader("CPIAUCSL", "fred", args.start, args.end)
    if cpi.empty:
        raise SystemExit("No CPI data returned from FRED.")
    cpi.index = pd.to_datetime(cpi.index)
    cpi = cpi.rename(columns={"CPIAUCSL": "CPI"})
    # Convert CPI to month-end timestamps for consistency
    cpi["Month"] = cpi.index.to_period("M").dt.to_timestamp("M")
    cpi_m = cpi.groupby("Month", as_index=False)["CPI"].last()

    # Merge prices with CPI
    df = price.merge(cpi_m, how="left", on="Month")
    # Forward-fill CPI in case some month-ends are missing
    df["CPI"] = df["CPI"].ffill()

    # Finalize export frame
    out = df[["Month", "AdjClose", "CPI"]].rename(columns={"Month": "Date"}).reset_index(drop=True)

    # Prepare Excel export
    outfile = args.outfile or f"{ticker}_inflation.xlsx"
    print(f"Writing Excel file to {outfile} ...", file=sys.stderr)

    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Data")

    # Reopen with openpyxl to add formulas and named ranges
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName

    wb = load_workbook(outfile)
    ws = wb["Data"]

    # Determine last row with data
    last_row = ws.max_row  # includes header row

    # Put CPI_END (last CPI value) in H1 and name it for clarity; optionally as formula to last row
    # We'll set H1 = value from last data row column C
    cpi_end_cell = f"C{last_row}"
    ws["H1"] = f"={cpi_end_cell}"
    ws["G1"] = "Notes:"
    ws["H2"] = "CPI_END (last CPI value)"
    # Create a defined name CPI_END pointing to H1
    dn = DefinedName(name="CPI_END", attr_text=f"Data!$H$1")
    wb.defined_names.append(dn)

    # Headers for helper columns
    ws["D1"] = "CPI Factor (=CPI_END/CPI_t)"
    ws["E1"] = "Real Price (=AdjClose*CPI Factor)"
    ws["F1"] = "Real % Change"
    # Fill formulas down from row 2 to last_row
    for r in range(2, last_row + 1):
        # D: CPI factor = CPI_END / C[row]
        ws[f"D{r}"] = f"=CPI_END/C{r}"
        # E: Real Price = B[row] * D[row]
        ws[f"E{r}"] = f"=B{r}*D{r}"
        # F: Real % Change vs previous row (leave F2 blank if no prior row)
        if r == 2:
            ws[f"F{r}"] = ""
        else:
            ws[f"F{r}"] = f"=IF(E{r-1}=0,\"\",E{r}/E{r-1}-1)"

    # Format columns with reasonable widths
    widths = {
        "A": 12,  # Date
        "B": 14,  # AdjClose
        "C": 10,  # CPI
        "D": 16,  # CPI Factor
        "E": 16,  # Real Price
        "F": 14,  # Real % Change
        "G": 10,
        "H": 24,  # CPI_END
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(outfile)
    print(f"Done. Wrote {outfile}", file=sys.stderr)

if __name__ == "__main__":
    main()
